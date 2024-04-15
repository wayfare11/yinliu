import pandas as pd
from openpyxl import load_workbook

# 创建数据
data = {
    "A": ["A1", "A2", "A3", "A4", "A5"],
    "B": ["A1", "B2", "B3", "B4", "B5"],
    "C": ["A1", "C2", "C3", "C4", "C5"],
    "D": ["A1", "D2", "D3", "D4", "D5"],
    "E": ["A1", "E2", "E3", "E4", "E5"],
    "F": ["A1", "F2", "F3", "F4", "F5"],
    "G": ["A1", "G2", "G3", "G4", "G5"],
    "H": ["A1", "H2", "H3", "H4", "H5"],
    "I": ["A1", "I2", "I3", "I4", "I5"],
    "J": ["A1", "J2", "J3", "J4", "J5"],
    "K": ["A1", "K2", "K3", "K4", "K5"]
}

df = pd.DataFrame(data)

# 导出为 Excel 文件
df.to_excel("output.xlsx", index=False)

# 打开 Excel 文件
book = load_workbook("output.xlsx")
sheet = book.active

# 合并第一行的所有列
sheet.merge_cells('A3:K3')

# 保存更改
book.save("output.xlsx")
